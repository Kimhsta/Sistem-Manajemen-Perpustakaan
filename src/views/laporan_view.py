# views/laporan_view.py
import csv
from datetime import date
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from database.db import get_conn

STATUS_OPTIONS = ["Semua", "Aktif", "Sudah Kembali", "Terlambat"]

class LaporanView(tk.Frame):
    """
    Laporan transaksi pinjam/kembali:
    - Filter: rentang tanggal (loan_date), status, keyword (nim/nama/kode/judul)
    - Ringkasan: total transaksi, aktif, terlambat, total denda
    - Ekspor: CSV (selalu) / Excel (opsional, jika openpyxl terpasang)
    """
    def __init__(self, app, **kwargs):
        super().__init__(app, **kwargs)
        self.app = app
        self._last_rows = []   # cache hasil query terakhir (list sqlite Row)
        self._build_ui()
        self._prefill_filters()
        self._reload()

    # ---------- UI ----------
    def _build_ui(self):
        self.columnconfigure(0, weight=1)

        # Header
        header = ttk.Frame(self, padding=12)
        header.grid(row=0, column=0, sticky="ew")
        ttk.Label(header, text="Laporan Transaksi & Denda", font=("Segoe UI", 16, "bold")).pack(side="left")
        ttk.Button(header, text="Kembali", command=self._go_back).pack(side="right")

        # Filter
        flt = ttk.LabelFrame(self, text="Filter", padding=12)
        flt.grid(row=1, column=0, sticky="ew", padx=12, pady=(0,8))
        for i in range(8):
            flt.columnconfigure(i, weight=1)

        ttk.Label(flt, text="Dari (YYYY-MM-DD)").grid(row=0, column=0, sticky="w", pady=3)
        self.ent_from = ttk.Entry(flt)
        self.ent_from.grid(row=0, column=1, sticky="ew", pady=3)

        ttk.Label(flt, text="Sampai (YYYY-MM-DD)").grid(row=0, column=2, sticky="w", pady=3)
        self.ent_to = ttk.Entry(flt)
        self.ent_to.grid(row=0, column=3, sticky="ew", pady=3)

        ttk.Label(flt, text="Status").grid(row=0, column=4, sticky="w", pady=3)
        self.cmb_status = ttk.Combobox(flt, values=STATUS_OPTIONS, state="readonly")
        self.cmb_status.grid(row=0, column=5, sticky="ew", pady=3)
        self.cmb_status.current(0)

        ttk.Label(flt, text="Kata kunci").grid(row=0, column=6, sticky="w", pady=3)
        self.ent_kw = ttk.Entry(flt)
        self.ent_kw.grid(row=0, column=7, sticky="ew", pady=3)

        ttk.Button(flt, text="Tampilkan", command=self._reload).grid(row=1, column=7, sticky="e", pady=4)
        ttk.Button(flt, text="Reset", command=self._reset_filters).grid(row=1, column=6, sticky="e", padx=6, pady=4)

        # Ringkasan
        sumf = ttk.LabelFrame(self, text="Ringkasan", padding=12)
        sumf.grid(row=2, column=0, sticky="ew", padx=12, pady=(0,8))
        for i in range(4):
            sumf.columnconfigure(i, weight=1)

        self.lbl_total = ttk.Label(sumf, text="Total Transaksi: 0")
        self.lbl_total.grid(row=0, column=0, sticky="w")
        self.lbl_aktif = ttk.Label(sumf, text="Aktif: 0")
        self.lbl_aktif.grid(row=0, column=1, sticky="w")
        self.lbl_telat = ttk.Label(sumf, text="Terlambat: 0")
        self.lbl_telat.grid(row=0, column=2, sticky="w")
        self.lbl_denda = ttk.Label(sumf, text="Total Denda: Rp0")
        self.lbl_denda.grid(row=0, column=3, sticky="w")

        # Tabel
        box = ttk.LabelFrame(self, text="Data Transaksi", padding=12)
        box.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0,8))
        self.rowconfigure(3, weight=1)
        box.rowconfigure(0, weight=1)
        box.columnconfigure(0, weight=1)

        cols = ("id","nim","nama","kode_buku","judul","qty","loan_date","due_date","return_date","total_fine")
        self.tree = ttk.Treeview(box, columns=cols, show="headings", height=14)
        widths = (60,120,160,100,220,60,100,100,100,90)
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c.upper())
            self.tree.column(c, width=w, anchor="center")
        self.tree.grid(row=0, column=0, sticky="nsew")

        y = ttk.Scrollbar(box, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=y.set)
        y.grid(row=0, column=1, sticky="ns")

        # Ekspor
        exp = ttk.Frame(self, padding=12)
        exp.grid(row=4, column=0, sticky="ew", padx=12, pady=(0,12))
        ttk.Button(exp, text="Ekspor CSV", command=self._export_csv).pack(side="left")
        ttk.Button(exp, text="Ekspor Excel (xlsx)", command=self._export_excel).pack(side="left", padx=8)

    # ---------- Helpers ----------
    def _go_back(self):
        from views.admin_view import AdminView
        self.app.show(AdminView)

    def _prefill_filters(self):
        # default: bulan berjalan
        today = date.today()
        first = today.replace(day=1).isoformat()
        self.ent_from.insert(0, first)
        self.ent_to.insert(0, today.isoformat())

    def _reset_filters(self):
        self.ent_from.delete(0, "end")
        self.ent_to.delete(0, "end")
        self.ent_kw.delete(0, "end")
        self.cmb_status.set("Semua")
        self._prefill_filters()
        self._reload()

    # ---------- Query & Data ----------
    def _reload(self):
        frm = self.ent_from.get().strip()
        to  = self.ent_to.get().strip()
        kw  = f"%{self.ent_kw.get().strip()}%"
        status = self.cmb_status.get()

        # Build WHERE dinamis
        where = ["DATE(l.loan_date) >= DATE(?)", "DATE(l.loan_date) <= DATE(?)"]
        params = [frm, to]

        if status == "Aktif":
            where.append("l.return_date IS NULL")
        elif status == "Sudah Kembali":
            where.append("l.return_date IS NOT NULL")
        elif status == "Terlambat":
            where.append("l.return_date IS NULL AND DATE(l.due_date) < DATE('now')")

        if kw != "%%":
            where.append("(s.nim LIKE ? OR s.nama LIKE ? OR b.kode LIKE ? OR b.judul LIKE ?)")
            params.extend([kw, kw, kw, kw])

        where_sql = " AND ".join(where)

        sql = f"""
        SELECT l.id, s.nim, s.nama, b.kode AS kode_buku, b.judul,
               l.qty, l.loan_date, l.due_date, l.return_date, l.total_fine
        FROM loans l
        JOIN students s ON s.id = l.student_id
        JOIN books b    ON b.id = l.book_id
        WHERE {where_sql}
        ORDER BY l.id DESC
        """

        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(sql, params)
            rows = cur.fetchall()

            # Ringkasan
            # total
            total = len(rows)
            # aktif
            cur.execute(f"SELECT COUNT(*) AS n FROM ({sql}) AS q WHERE q.return_date IS NULL", params)
            aktif = (cur.fetchone() or {"n": 0})["n"]
            # terlambat
            cur.execute(f"SELECT COUNT(*) AS n FROM ({sql}) AS q WHERE q.return_date IS NULL AND DATE(q.due_date) < DATE('now')", params)
            telat = (cur.fetchone() or {"n": 0})["n"]
            # total denda (yang sudah kembali)
            cur.execute(f"SELECT COALESCE(SUM(q.total_fine),0) AS tot FROM ({sql}) AS q WHERE q.return_date IS NOT NULL", params)
            tot_denda = (cur.fetchone() or {"tot": 0})["tot"]

        # render
        self._last_rows = rows
        self.tree.delete(*self.tree.get_children())
        for r in rows:
            self.tree.insert("", "end", values=(
                r["id"], r["nim"], r["nama"], r["kode_buku"], r["judul"], r["qty"],
                r["loan_date"], r["due_date"], r["return_date"] or "", r["total_fine"]
            ))

        self.lbl_total.config(text=f"Total Transaksi: {total}")
        self.lbl_aktif.config(text=f"Aktif: {aktif}")
        self.lbl_telat.config(text=f"Terlambat: {telat}")
        self.lbl_denda.config(text=f"Total Denda: Rp{tot_denda}")

    # ---------- Export ----------
    def _export_csv(self):
        if not self._last_rows:
            messagebox.showwarning("Perhatian", "Tidak ada data untuk diekspor.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="laporan_perpustakaan.csv"
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["id","nim","nama","kode_buku","judul","qty","loan_date","due_date","return_date","total_fine"])
                for r in self._last_rows:
                    w.writerow([
                        r["id"], r["nim"], r["nama"], r["kode_buku"], r["judul"], r["qty"],
                        r["loan_date"], r["due_date"], r["return_date"] or "", r["total_fine"]
                    ])
            messagebox.showinfo("Sukses", f"CSV disimpan ke:\n{path}")
        except Exception as e:
            messagebox.showerror("Gagal", str(e))

    def _export_excel(self):
        if not self._last_rows:
            messagebox.showwarning("Perhatian", "Tidak ada data untuk diekspor.")
            return
        try:
            import openpyxl  # optional
            from openpyxl.utils import get_column_letter
        except Exception:
            messagebox.showwarning(
                "Info",
                "openpyxl belum terpasang. Menggunakan ekspor CSV sebagai alternatif."
            )
            self._export_csv()
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile="laporan_perpustakaan.xlsx"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Laporan"

            headers = ["ID","NIM","Nama","Kode Buku","Judul","Qty","Loan Date","Due Date","Return Date","Total Fine"]
            ws.append(headers)

            for r in self._last_rows:
                ws.append([
                    r["id"], r["nim"], r["nama"], r["kode_buku"], r["judul"], r["qty"],
                    r["loan_date"], r["due_date"], r["return_date"] or "", r["total_fine"]
                ])

            # auto width sederhana
            for col in range(1, len(headers)+1):
                ws.column_dimensions[get_column_letter(col)].width = 16

            wb.save(path)
            messagebox.showinfo("Sukses", f"Excel disimpan ke:\n{path}")
        except Exception as e:
            messagebox.showerror("Gagal", str(e))
